import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
import database  
import time

print("enter the ending readings separated by commas:")
for end in database.beginning_readings.keys():
    e = float(input(f"Ending reading for flat {end}: "))
    database.ending_readings[end] = int(e)

unit_price = (input("enter the unit price: "))
month = input("enter the month: ")
year = input("enter the year: ")

# Converting d ictionary values to lists
balancel = list(database.balance.values())
beg_redl = list(database.beginning_readings.values())
end_redl = list(database.ending_readings.values())
flat_no  =  list(database.beginning_readings.keys())
unit_no = [unit_price]*len(flat_no)

# Calculating total units consumed
total_units = []
for i,j in zip(beg_redl,end_redl):
    total_units.append(j - i)

#this calculates the amount per flat and round off
amount = []
for k in total_units:
    amount.append(round(k*unit_no))

# calculating total amount including balance
total_amt = []
for o,n in zip(amount,balancel):
    total_amt.append(o + n)    

# calculating totals of units of apt
total_unitsk = 0
for l in total_units:
    total_unitsk += l  

# calculating totals of amount of apt
total_amth = 0
for d in amount:
    total_amth += d 

# calculating totals of total amount of apt with balance
total_amtr = 0
for f in total_amt:
    total_amtr += f

# Creating a DataFrame and exporting to Excel
data = {
    
"flat no" :  flat_no,

"beggnig reading" : beg_redl,

"ending reading" : end_redl,

"Amount per unit" : unit_no,

"total units" : total_units,

"sep bills" : amount,

"balance" : balancel,

"toatl amount" : total_amt

}

file = pd.DataFrame(data)

file_name = f"Water bill {month}:{year}.xlsx"
file.to_excel(file_name, index=False)

# Formatting the Excel file

wb = load_workbook(file_name)
ws = wb.active

#spacing and aligning
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    col_letter = column_cells[0].column_letter
    ws.column_dimensions[col_letter].width = length + 2

wb.save(file_name)

#exchange
database.ending_readings = database.beginning_readings.copy()
database.ending_readings.clear()
no_of = 1
# Final touches and saving
print("Formatting the Excel table...")
time.sleep(2)
print("spacing")
time.sleep(1)
print("in 1..")
time.sleep(1)
print("✅ Excel table created successfully as 'water bill.xlsx'")

